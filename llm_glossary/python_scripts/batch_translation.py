import csv
import os
import re
import sys
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from evaluation.metrics import evaluate_candidates
from evaluation.reporting import format_evaluation_summary


API_KEY = os.getenv("API_KEY")
BASE_URL = os.getenv("BASE_URL", "https://openwebui-test.hrz.uni-bonn.de/api")
MODEL = os.getenv("MODEL_NAME", "local.openai/gpt-oss-120b")
GLOSSARY_CSV = os.getenv("GLOSSARY_CSV", "data/glossary.csv")
OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "outputs/translation_results.xlsx")
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "60"))
MAX_SAMPLES = int(os.getenv("MAX_SAMPLES", "2000"))


def is_sample(row):
    # Keep only useful text samples.
    de_text = row["de"].strip()
    en_text = row["en"].strip()
    if not de_text or not en_text:
        return False
    if "http" in de_text.lower() or "www." in de_text.lower():
        return False
    return len(de_text.split()) >= 8 and len(en_text.split()) >= 8


def read_samples():
    # Read translation examples from glossary.
    samples = []
    with open(GLOSSARY_CSV, encoding="utf-8") as file:
        for row in csv.DictReader(file):
            if is_sample(row):
                samples.append((row["de"].strip(), row["en"].strip()))
            if MAX_SAMPLES and len(samples) == MAX_SAMPLES:
                break
    return samples


def find_glossary_terms(german_text):
    # Find matching glossary term entries.
    matches = []
    with open(GLOSSARY_CSV, encoding="utf-8") as file:
        for row in csv.DictReader(file):
            de_text = row["de"].strip()
            en_text = row["en"].strip()
            if de_text == german_text:
                continue
            pattern = r"\b" + re.escape(de_text.lower()) + r"\b"
            if re.search(pattern, german_text.lower()):
                matches.append(f"{de_text} = {en_text}")
            if len(matches) == 8:
                break
    return "\n".join(matches)


def translate(german_text, glossary=""):
    # Build the prompt for translation.
    if glossary:
        prompt = f"""Translate this German text into English.
Use the glossary if relevant.
Return only the English translation. Do not use Markdown or explanations.

Glossary:
{glossary}

German:
{german_text}

English:"""
    else:
        prompt = f"""Translate this German text into English.
Return only the English translation. Do not use Markdown or explanations.

German:
{german_text}

English:"""

    # Send translation request to LLM.
    response = requests.post(
        f"{BASE_URL}/chat/completions",
        headers={
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json",
        },
        json={
            "model": MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0,
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    return response.json()["choices"][0]["message"]["content"].strip()


def write_excel(rows, without_metrics: dict, with_metrics: dict):
    # Create workbook and result sheets.
    workbook = Workbook()
    translations_sheet = workbook.active
    translations_sheet.title = "translations"

    headers = [
        "German",
        "Reference",
        "Without Glossary",
        "With Glossary",
        "BLEU (No Glossary)",
        "BLEU (Glossary)",
        "COMET (No Glossary)",
        "COMET (Glossary)",
        "BERTScore Precision (No Glossary)",
        "BERTScore Precision (Glossary)",
        "BERTScore Recall (No Glossary)",
        "BERTScore Recall (Glossary)",
        "BERTScore F1 (No Glossary)",
        "BERTScore F1 (Glossary)",
    ]
    translations_sheet.append(headers)

    # Write one result row per sentence.
    for row, without_score, with_score in zip(
        rows, without_metrics["sentence_scores"], with_metrics["sentence_scores"]
    ):
        translations_sheet.append(
            [
                row["german_text"],
                row["ground_truth"],
                row["without_glossary"],
                row["with_glossary"],
                without_score["sacrebleu"],
                with_score["sacrebleu"],
                without_score["comet"],
                with_score["comet"],
                without_score["bertscore_precision"],
                with_score["bertscore_precision"],
                without_score["bertscore_recall"],
                with_score["bertscore_recall"],
                without_score["bertscore_f1"],
                with_score["bertscore_f1"],
            ]
        )

    # Improve final spreadsheet readability.
    translations_sheet.freeze_panes = "A2"
    widths = [55, 55, 55, 55, 18, 18, 18, 18, 24, 24, 22, 22, 22, 22]
    for index, width in enumerate(widths, start=1):
        translations_sheet.column_dimensions[_excel_column(index)].width = width

    for cell in translations_sheet[1]:
        cell.font = Font(bold=True)

    for row in translations_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in translations_sheet.iter_rows(min_row=2, min_col=5, max_col=14):
        for cell in row:
            cell.number_format = "0.0000"

    # Add the summary metric sheet.
    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.append(["System", "Metric", "Value"])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

    summary_rows = [
        (
            "Without Glossary",
            "Average sentence SacreBLEU",
            without_metrics["average_sacrebleu"],
        ),
        ("Without Glossary", "Corpus SacreBLEU", without_metrics["corpus_sacrebleu"]),
        ("Without Glossary", "Average COMET", without_metrics["average_comet"]),
        (
            "Without Glossary",
            "Average BERTScore Precision",
            without_metrics["average_bertscore_precision"],
        ),
        (
            "Without Glossary",
            "Average BERTScore Recall",
            without_metrics["average_bertscore_recall"],
        ),
        (
            "Without Glossary",
            "Average BERTScore F1",
            without_metrics["average_bertscore_f1"],
        ),
        (
            "With Glossary",
            "Average sentence SacreBLEU",
            with_metrics["average_sacrebleu"],
        ),
        ("With Glossary", "Corpus SacreBLEU", with_metrics["corpus_sacrebleu"]),
        ("With Glossary", "Average COMET", with_metrics["average_comet"]),
        (
            "With Glossary",
            "Average BERTScore Precision",
            with_metrics["average_bertscore_precision"],
        ),
        (
            "With Glossary",
            "Average BERTScore Recall",
            with_metrics["average_bertscore_recall"],
        ),
        ("With Glossary", "Average BERTScore F1", with_metrics["average_bertscore_f1"]),
    ]
    for summary_row in summary_rows:
        summary_sheet.append(summary_row)
    summary_sheet.column_dimensions["A"].width = 22
    summary_sheet.column_dimensions["B"].width = 32
    summary_sheet.column_dimensions["C"].width = 16
    for cell in summary_sheet["C"][1:]:
        cell.number_format = "0.0000"

    output = Path(OUTPUT_XLSX)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def _excel_column(index):
    # Convert number into Excel column.
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


if not API_KEY:
    raise SystemExit("Set API_KEY first.")


# Store all translation result dictionaries.
translations = []
samples = read_samples()

for index, (german_text, ground_truth) in enumerate(samples, start=1):
    # Translate each selected text sample.
    glossary = find_glossary_terms(german_text)
    without_glossary = translate(german_text)
    with_glossary = translate(german_text, glossary)

    translations.append(
        {
            "german_text": german_text,
            "ground_truth": ground_truth,
            "without_glossary": without_glossary,
            "with_glossary": with_glossary,
        }
    )

    print(f"Translated {index}/{len(samples)}", flush=True)

# Prepare aligned lists for evaluation.
sources = [row["german_text"] for row in translations]
references = [row["ground_truth"] for row in translations]
without_hypotheses = [row["without_glossary"] for row in translations]
with_hypotheses = [row["with_glossary"] for row in translations]

print("Evaluating translations without glossary...", flush=True)
without_metrics = evaluate_candidates(sources, references, without_hypotheses)

print("Evaluating translations with glossary...", flush=True)
with_metrics = evaluate_candidates(sources, references, with_hypotheses)

print(format_evaluation_summary(without_metrics, with_metrics), flush=True)

output_path = write_excel(translations, without_metrics, with_metrics)
print(f"Wrote {output_path}")
